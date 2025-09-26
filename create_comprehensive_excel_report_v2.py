#!/usr/bin/env python3
"""
HVDC 종합 엑셀 리포트 생성 시스템 v2.0.0
- 3개 시트 구조: 전체_트랜잭션_데이터, 창고_월별_입출고, 현장_월별_입고재고
- SQM, Stack_Status, Flow Code 분석 포함
- 전문적 엑셀 포맷팅 및 차트
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from openpyxl.utils import get_column_letter

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ComprehensiveExcelReportGenerator:
    """종합 엑셀 리포트 생성기 v2.0"""
    
    def __init__(self):
        """시스템 초기화"""
        self.output_file = f"HVDC_종합_엑셀_리포트_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.data_file = "HVDC_DHL_Warehouse_완전복구_20250704_125724.xlsx"
        
        # 데이터 로드
        self.load_data()
        
        logger.info("종합 엑셀 리포트 생성 시스템 v2.0 초기화 완료")
    
    def load_data(self):
        """데이터 로드 및 전처리"""
        try:
            logger.info(f"데이터 파일 로드: {self.data_file}")
            
            # 원본 데이터 로드
            self.df = pd.read_excel(self.data_file)
            logger.info(f"원본 데이터 로드 완료: {len(self.df)}건")
            
            # 날짜 컬럼 변환
            date_columns = [col for col in self.df.columns if 'Date' in col or 'date' in col]
            for col in date_columns:
                self.df[col] = pd.to_datetime(self.df[col], errors='coerce')
            
            # 숫자 컬럼 안전 변환
            numeric_columns = ['Pkg', 'CBM', 'N.W', 'G.W']
            for col in numeric_columns:
                if col in self.df.columns:
                    self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0)
            
            # Flow Code 분석
            self.analyze_flow_codes()
            
            # SQM 및 Stack_Status 확인
            self.check_sqm_stack_status()
            
            logger.info("데이터 전처리 완료")
            
        except Exception as e:
            logger.error(f"데이터 로드 실패: {e}")
            raise
    
    def analyze_flow_codes(self):
        """Flow Code 분석"""
        if 'Flow_Code' in self.df.columns:
            flow_code_counts = self.df['Flow_Code'].value_counts().sort_index()
            logger.info(f"Flow Code 분포: {dict(flow_code_counts)}")
            
            # Flow Code 패턴 분석
            self.flow_code_patterns = {
                'Flow_0': 'Port → Site 직접 (창고 경유 없음)',
                'Flow_1': '창고 1개 경유',
                'Flow_2': '창고 2개 경유', 
                'Flow_3': '창고 3개+ 경유'
            }
    
    def check_sqm_stack_status(self):
        """SQM 및 Stack_Status 컬럼 확인"""
        self.has_sqm = 'SQM' in self.df.columns
        self.has_stack_status = 'Stack_Status' in self.df.columns
        
        logger.info(f"SQM 컬럼 존재: {self.has_sqm}")
        logger.info(f"Stack_Status 컬럼 존재: {self.has_stack_status}")
        
        if self.has_sqm:
            sqm_counts = self.df['SQM'].value_counts()
            logger.info(f"SQM 분포: {dict(sqm_counts.head())}")
        
        if self.has_stack_status:
            stack_counts = self.df['Stack_Status'].value_counts()
            logger.info(f"Stack_Status 분포: {dict(stack_counts.head())}")
    
    def adjust_column_widths(self, ws, start_row=1, end_row=None):
        """컬럼 너비 자동 조정 (MergedCell 오류 방지)"""
        if end_row is None:
            end_row = ws.max_row
        
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for row_num in range(start_row, end_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_transaction_data_sheet(self, wb):
        """전체_트랜잭션_데이터 시트 생성"""
        ws = wb.create_sheet("전체_트랜잭션_데이터")
        
        # 제목
        ws['A1'] = "HVDC 전체 트랜잭션 데이터 (7,573건)"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:Z1')
        
        # 데이터 통계
        ws['A3'] = "📊 데이터 통계"
        ws['A3'].font = Font(size=14, bold=True, color="366092")
        
        stats_data = [
            ["총 트랜잭션", f"{len(self.df):,}건"],
            ["컬럼 수", f"{len(self.df.columns)}개"],
            ["SQM 포함", "✅" if self.has_sqm else "❌"],
            ["Stack_Status 포함", "✅" if self.has_stack_status else "❌"],
            ["Flow Code 포함", "✅" if 'Flow_Code' in self.df.columns else "❌"]
        ]
        
        for i, (key, value) in enumerate(stats_data):
            ws[f'A{i+5}'] = key
            ws[f'B{i+5}'] = value
            ws[f'A{i+5}'].font = Font(bold=True)
            ws[f'B{i+5}'].font = Font(color="366092")
        
        # Flow Code 분석 (있는 경우)
        if 'Flow_Code' in self.df.columns:
            ws['A11'] = "🔄 Flow Code 분석"
            ws['A11'].font = Font(size=14, bold=True, color="366092")
            
            flow_data = [["Flow Code", "설명", "건수"]]
            for code, description in self.flow_code_patterns.items():
                count = len(self.df[self.df['Flow_Code'] == int(code.split('_')[1])])
                flow_data.append([code, description, f"{count:,}건"])
            
            for i, row in enumerate(flow_data):
                for j, value in enumerate(row):
                    cell = ws.cell(row=13+i, column=1+j, value=value)
                    if i == 0:  # 헤더
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 데이터 시작 위치
        data_start_row = 20
        
        # 컬럼 헤더
        for i, col in enumerate(self.df.columns):
            cell = ws.cell(row=data_start_row, column=i+1, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 데이터 입력
        for i, row in enumerate(dataframe_to_rows(self.df, index=False, header=False)):
            for j, value in enumerate(row):
                cell = ws.cell(row=data_start_row+1+i, column=j+1, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0'
        
        # 컬럼 너비 조정
        self.adjust_column_widths(ws, data_start_row, data_start_row + len(self.df))
        
        logger.info("전체_트랜잭션_데이터 시트 생성 완료")
    
    def create_warehouse_monthly_sheet(self, wb):
        """창고_월별_입출고 시트 생성"""
        ws = wb.create_sheet("창고_월별_입출고")
        
        # 제목
        ws['A1'] = "창고별 월별 입고/출고 현황"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:Z1')
        
        # 창고별 월별 집계
        if 'Warehouse' in self.df.columns and 'In_Date' in self.df.columns:
            # 입고 데이터 집계
            self.df['In_Month'] = self.df['In_Date'].dt.to_period('M')
            warehouse_inbound = self.df.groupby(['Warehouse', 'In_Month']).agg({
                'Pkg': 'sum',
                'CBM': 'sum',
                'N.W': 'sum',
                'G.W': 'sum'
            }).reset_index()
            
            # 출고 데이터 집계 (Out_Date가 있는 경우)
            if 'Out_Date' in self.df.columns:
                self.df['Out_Month'] = self.df['Out_Date'].dt.to_period('M')
                warehouse_outbound = self.df.groupby(['Warehouse', 'Out_Month']).agg({
                    'Pkg': 'sum',
                    'CBM': 'sum',
                    'N.W': 'sum',
                    'G.W': 'sum'
                }).reset_index()
                warehouse_outbound = warehouse_outbound.rename(columns={
                    'Out_Month': 'Month',
                    'Pkg': 'Out_Pkg',
                    'CBM': 'Out_CBM',
                    'N.W': 'Out_NW',
                    'G.W': 'Out_GW'
                })
            else:
                warehouse_outbound = pd.DataFrame()
            
            # 입고 데이터 헤더
            ws['A3'] = "🏭 창고별 월별 입고 현황"
            ws['A3'].font = Font(size=14, bold=True, color="366092")
            
            # 입고 데이터 헤더
            headers = ["창고명", "월", "입고_Pkg", "입고_CBM", "입고_N.W", "입고_G.W"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=5, column=i+1, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 입고 데이터 입력
            for i, row in warehouse_inbound.iterrows():
                for j, value in enumerate(row):
                    cell = ws.cell(row=6+i, column=j+1, value=value)
                    if isinstance(value, (int, float)) and j > 1:  # 숫자 컬럼
                        cell.number_format = '#,##0'
            
            # 출고 데이터 (있는 경우)
            if not warehouse_outbound.empty:
                ws['H3'] = "📤 창고별 월별 출고 현황"
                ws['H3'].font = Font(size=14, bold=True, color="366092")
                
                # 출고 데이터 헤더
                out_headers = ["창고명", "월", "출고_Pkg", "출고_CBM", "출고_N.W", "출고_G.W"]
                for i, header in enumerate(out_headers):
                    cell = ws.cell(row=5, column=i+8, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 출고 데이터 입력
                for i, row in warehouse_outbound.iterrows():
                    for j, value in enumerate(row):
                        cell = ws.cell(row=6+i, column=j+8, value=value)
                        if isinstance(value, (int, float)) and j > 1:  # 숫자 컬럼
                            cell.number_format = '#,##0'
        
        # 컬럼 너비 조정
        self.adjust_column_widths(ws, 1, ws.max_row)
        
        logger.info("창고_월별_입출고 시트 생성 완료")
    
    def create_site_monthly_sheet(self, wb):
        """현장_월별_입고재고 시트 생성"""
        ws = wb.create_sheet("현장_월별_입고재고")
        
        # 제목
        ws['A1'] = "현장별 월별 입고/재고 현황"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:Z1')
        
        # 현장별 월별 집계
        if 'Site' in self.df.columns and 'In_Date' in self.df.columns:
            # 입고 데이터 집계
            self.df['In_Month'] = self.df['In_Date'].dt.to_period('M')
            site_inbound = self.df.groupby(['Site', 'In_Month']).agg({
                'Pkg': 'sum',
                'CBM': 'sum',
                'N.W': 'sum',
                'G.W': 'sum'
            }).reset_index()
            
            # 누적재고 계산 (현장 특성: 출고 없음)
            site_inbound = site_inbound.sort_values(['Site', 'In_Month'])
            site_inbound['누적_Pkg'] = site_inbound.groupby('Site')['Pkg'].cumsum()
            site_inbound['누적_CBM'] = site_inbound.groupby('Site')['CBM'].cumsum()
            site_inbound['누적_NW'] = site_inbound.groupby('Site')['N.W'].cumsum()
            site_inbound['누적_GW'] = site_inbound.groupby('Site')['G.W'].cumsum()
            
            # 현장 정보
            ws['A3'] = "🏗️ 현장별 월별 입고 현황"
            ws['A3'].font = Font(size=14, bold=True, color="366092")
            
            # 현장 특성 설명
            ws['A4'] = "현장 특성: 출고 없음 (누적재고만 관리)"
            ws['A4'].font = Font(italic=True, color="FF6600")
            
            # 입고 데이터 헤더
            headers = ["현장명", "월", "입고_Pkg", "입고_CBM", "입고_N.W", "입고_G.W"]
            for i, header in enumerate(headers):
                cell = ws.cell(row=6, column=i+1, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 입고 데이터 입력
            for i, row in site_inbound.iterrows():
                for j, value in enumerate(row[:6]):  # 입고 데이터만
                    cell = ws.cell(row=7+i, column=j+1, value=value)
                    if isinstance(value, (int, float)) and j > 1:  # 숫자 컬럼
                        cell.number_format = '#,##0'
            
            # 누적재고 데이터
            ws['H3'] = "📦 현장별 월별 누적재고 현황"
            ws['H3'].font = Font(size=14, bold=True, color="366092")
            
            # 누적재고 헤더
            cum_headers = ["현장명", "월", "누적_Pkg", "누적_CBM", "누적_N.W", "누적_G.W"]
            for i, header in enumerate(cum_headers):
                cell = ws.cell(row=6, column=i+8, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 누적재고 데이터 입력
            for i, row in site_inbound.iterrows():
                for j, value in enumerate(row[['Site', 'In_Month', '누적_Pkg', '누적_CBM', '누적_NW', '누적_GW']]):
                    cell = ws.cell(row=7+i, column=j+8, value=value)
                    if isinstance(value, (int, float)) and j > 1:  # 숫자 컬럼
                        cell.number_format = '#,##0'
        
        # 컬럼 너비 조정
        self.adjust_column_widths(ws, 1, ws.max_row)
        
        logger.info("현장_월별_입고재고 시트 생성 완료")
    
    def generate_excel_report(self):
        """엑셀 리포트 생성"""
        logger.info("종합 엑셀 리포트 생성 시작")
        
        try:
            # 워크북 생성
            wb = Workbook()
            
            # 기본 시트 제거
            wb.remove(wb.active)
            
            # 시트 생성
            self.create_transaction_data_sheet(wb)
            self.create_warehouse_monthly_sheet(wb)
            self.create_site_monthly_sheet(wb)
            
            # 파일 저장
            wb.save(self.output_file)
            logger.info(f"종합 엑셀 리포트 생성 완료: {self.output_file}")
            
            # 생성 결과 리포트
            report = {
                "생성_시간": datetime.now().isoformat(),
                "파일명": self.output_file,
                "시트_수": len(wb.sheetnames),
                "시트_구성": wb.sheetnames,
                "총_트랜잭션": len(self.df),
                "컬럼_수": len(self.df.columns),
                "SQM_포함": self.has_sqm,
                "Stack_Status_포함": self.has_stack_status,
                "Flow_Code_포함": 'Flow_Code' in self.df.columns
            }
            
            # 리포트 저장
            report_file = f"종합_엑셀_리포트_생성_리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            
            logger.info(f"생성 리포트 저장: {report_file}")
            
            return report
            
        except Exception as e:
            logger.error(f"엑셀 리포트 생성 실패: {e}")
            raise

def main():
    """메인 실행 함수"""
    print("📊 HVDC 종합 엑셀 리포트 생성 시스템 v2.0")
    print("=" * 60)
    
    try:
        generator = ComprehensiveExcelReportGenerator()
        report = generator.generate_excel_report()
        
        print(f"\n✅ 종합 엑셀 리포트 생성 완료!")
        print(f"📁 파일: {report['파일명']}")
        print(f"📋 시트 수: {report['시트_수']}개")
        print(f"📊 시트 구성: {', '.join(report['시트_구성'])}")
        print(f"📈 총 트랜잭션: {report['총_트랜잭션']:,}건")
        print(f"📋 컬럼 수: {report['컬럼_수']}개")
        print(f"🔍 SQM 포함: {'✅' if report['SQM_포함'] else '❌'}")
        print(f"📦 Stack_Status 포함: {'✅' if report['Stack_Status_포함'] else '❌'}")
        print(f"🔄 Flow_Code 포함: {'✅' if report['Flow_Code_포함'] else '❌'}")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        return False
    
    return True

if __name__ == "__main__":
    main() 