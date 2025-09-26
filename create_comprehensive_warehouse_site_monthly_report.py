#!/usr/bin/env python3
"""
HVDC 종합 월별 보고서 생성 시스템 v1.0.0
- 전체 월별 창고별 입고/출고 시트
- 전체 월별 현장별 입고/누적재고 시트
- TDD 방식으로 구현
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ComprehensiveMonthlyReportGenerator:
    """종합 월별 보고서 생성기"""
    
    def __init__(self):
        """시스템 초기화"""
        self.data_file = "HVDC_DHL_Warehouse_완전복구_20250704_125724.xlsx"
        self.output_file = f"HVDC_종합_월별_보고서_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 창고 목록
        self.warehouses = [
            'DSV Indoor', 'DSV Al Markaz', 'DSV Outdoor', 'AAA  Storage',
            'Hauler Indoor', 'DSV MZP', 'MOSB', 'DHL Warehouse'
        ]
        
        # 현장 목록
        self.sites = [
            'Shifting', 'MIR', 'SHU', 'DAS', 'AGI'
        ]
        
        logger.info("종합 월별 보고서 생성 시스템 초기화 완료")
    
    def load_data(self):
        """데이터 로드 및 전처리"""
        logger.info("데이터 로드 시작")
        
        try:
            df = pd.read_excel(self.data_file)
            logger.info(f"데이터 로드 완료: {len(df)}개 레코드, {len(df.columns)}개 컬럼")
            
            # 날짜 컬럼 처리
            date_columns = ['ETD/ATD', 'ETA/ATA', 'DSV Indoor', 'DSV Al Markaz', 
                           'DSV Outdoor', 'AAA  Storage', 'Hauler Indoor', 
                           'DSV MZP', 'MOSB', 'Shifting', 'MIR', 'SHU', 'DAS', 
                           'AGI', 'DHL Warehouse']
            
            for col in date_columns:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            # 숫자 컬럼 처리
            numeric_columns = ['Pkg', 'CBM', 'N.W(kgs)', 'G.W(kgs)']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # 월별 데이터 생성 (ETD/ATD가 없는 경우 다른 날짜 컬럼 사용)
            if 'ETD/ATD' in df.columns and df['ETD/ATD'].notna().any():
                df['Year'] = df['ETD/ATD'].dt.year
                df['Month'] = df['ETD/ATD'].dt.month
                df['YearMonth'] = df['ETD/ATD'].dt.to_period('M')
            else:
                # 첫 번째 유효한 날짜 컬럼 사용
                for col in date_columns:
                    if col in df.columns and df[col].notna().any():
                        df['Year'] = df[col].dt.year
                        df['Month'] = df[col].dt.month
                        df['YearMonth'] = df[col].dt.to_period('M')
                        break
                else:
                    # 날짜가 없는 경우 기본값 설정
                    df['Year'] = 2024
                    df['Month'] = 1
                    df['YearMonth'] = pd.Period('2024-01')
            
            return df
            
        except Exception as e:
            logger.error(f"데이터 로드 실패: {e}")
            raise
    
    def create_warehouse_monthly_summary(self, df):
        """창고별 월별 입고/출고 요약 생성"""
        logger.info("창고별 월별 요약 생성 시작")
        
        warehouse_summary = []
        
        for warehouse in self.warehouses:
            if warehouse not in df.columns:
                continue
                
            # 입고 데이터 (해당 창고에 도착한 날짜)
            warehouse_data = df[df[warehouse].notna()].copy()
            
            if len(warehouse_data) == 0:
                continue
            
            # 숫자 컬럼 확인 및 변환
            numeric_cols = {}
            for col in ['Pkg', 'CBM', 'N.W(kgs)', 'G.W(kgs)']:
                if col in warehouse_data.columns:
                    warehouse_data[col] = pd.to_numeric(warehouse_data[col], errors='coerce').fillna(0)
                    numeric_cols[col] = col
            
            if not numeric_cols:
                continue
            
            # 월별 집계 (안전한 방식)
            agg_dict = {}
            for col in numeric_cols:
                agg_dict[col] = ['sum', 'count']
            
            monthly_summary = warehouse_data.groupby(['Year', 'Month']).agg(agg_dict).reset_index()
            
            # 컬럼명 정리
            new_columns = ['Year', 'Month']
            for col in numeric_cols:
                new_columns.extend([f'Total_{col}', f'Count_{col}'])
            
            monthly_summary.columns = new_columns
            
            # 창고명 추가
            monthly_summary['Warehouse'] = warehouse
            
            warehouse_summary.append(monthly_summary)
        
        if warehouse_summary:
            result = pd.concat(warehouse_summary, ignore_index=True)
            result = result.sort_values(['Year', 'Month', 'Warehouse'])
            
            # 출고 데이터 계산 (다음 창고로 이동)
            if 'Total_Pkg' in result.columns:
                result['Outbound'] = result.groupby(['Year', 'Month'])['Total_Pkg'].shift(-1).fillna(0)
            
            logger.info(f"창고별 월별 요약 완료: {len(result)}개 레코드")
            return result
        else:
            logger.warning("창고별 데이터가 없습니다")
            return pd.DataFrame()
    
    def create_site_monthly_summary(self, df):
        """현장별 월별 입고/누적재고 요약 생성"""
        logger.info("현장별 월별 요약 생성 시작")
        
        site_summary = []
        
        for site in self.sites:
            if site not in df.columns:
                continue
                
            # 현장 도착 데이터
            site_data = df[df[site].notna()].copy()
            
            if len(site_data) == 0:
                continue
            
            # 숫자 컬럼 확인 및 변환
            numeric_cols = {}
            for col in ['Pkg', 'CBM', 'N.W(kgs)', 'G.W(kgs)']:
                if col in site_data.columns:
                    site_data[col] = pd.to_numeric(site_data[col], errors='coerce').fillna(0)
                    numeric_cols[col] = col
            
            if not numeric_cols:
                continue
            
            # 월별 집계 (안전한 방식)
            agg_dict = {}
            for col in numeric_cols:
                agg_dict[col] = ['sum', 'count']
            
            monthly_summary = site_data.groupby(['Year', 'Month']).agg(agg_dict).reset_index()
            
            # 컬럼명 정리
            new_columns = ['Year', 'Month']
            for col in numeric_cols:
                new_columns.extend([f'Total_{col}', f'Count_{col}'])
            
            monthly_summary.columns = new_columns
            
            # 현장명 추가
            monthly_summary['Site'] = site
            
            site_summary.append(monthly_summary)
        
        if site_summary:
            result = pd.concat(site_summary, ignore_index=True)
            result = result.sort_values(['Year', 'Month', 'Site'])
            
            # 누적재고 계산
            for col in ['Pkg', 'CBM']:
                total_col = f'Total_{col}'
                if total_col in result.columns:
                    result[f'Cumulative_{col}'] = result.groupby('Site')[total_col].cumsum()
            
            logger.info(f"현장별 월별 요약 완료: {len(result)}개 레코드")
            return result
        else:
            logger.warning("현장별 데이터가 없습니다")
            return pd.DataFrame()
    
    def create_detailed_warehouse_report(self, df):
        """창고별 상세 입고/출고 보고서"""
        logger.info("창고별 상세 보고서 생성 시작")
        
        detailed_reports = []
        
        for warehouse in self.warehouses:
            if warehouse not in df.columns:
                continue
            
            # 해당 창고 데이터
            warehouse_data = df[df[warehouse].notna()].copy()
            
            if len(warehouse_data) == 0:
                continue
            
            # 필요한 컬럼만 선택
            columns = ['Case No.', 'HVDC CODE', 'Description', 'Pkg', 'CBM', 
                      'N.W(kgs)', 'G.W(kgs)', 'ETD/ATD', warehouse, 'Status_WAREHOUSE']
            
            available_columns = [col for col in columns if col in warehouse_data.columns]
            warehouse_data = warehouse_data[available_columns].copy()
            
            # 창고명 추가
            warehouse_data['Warehouse'] = warehouse
            
            # 월별로 그룹화
            warehouse_data['Year'] = warehouse_data[warehouse].dt.year
            warehouse_data['Month'] = warehouse_data[warehouse].dt.month
            
            detailed_reports.append(warehouse_data)
        
        if detailed_reports:
            result = pd.concat(detailed_reports, ignore_index=True)
            result = result.sort_values(['Warehouse', 'Year', 'Month', 'Case No.'])
            
            logger.info(f"창고별 상세 보고서 완료: {len(result)}개 레코드")
            return result
        else:
            logger.warning("창고별 상세 데이터가 없습니다")
            return pd.DataFrame()
    
    def create_detailed_site_report(self, df):
        """현장별 상세 입고/누적재고 보고서"""
        logger.info("현장별 상세 보고서 생성 시작")
        
        detailed_reports = []
        
        for site in self.sites:
            if site not in df.columns:
                continue
            
            # 해당 현장 데이터
            site_data = df[df[site].notna()].copy()
            
            if len(site_data) == 0:
                continue
            
            # 필요한 컬럼만 선택
            columns = ['Case No.', 'HVDC CODE', 'Description', 'Pkg', 'CBM', 
                      'N.W(kgs)', 'G.W(kgs)', 'ETD/ATD', site, 'Status_SITE']
            
            available_columns = [col for col in columns if col in site_data.columns]
            site_data = site_data[available_columns].copy()
            
            # 현장명 추가
            site_data['Site'] = site
            
            # 월별로 그룹화
            site_data['Year'] = site_data[site].dt.year
            site_data['Month'] = site_data[site].dt.month
            
            detailed_reports.append(site_data)
        
        if detailed_reports:
            result = pd.concat(detailed_reports, ignore_index=True)
            result = result.sort_values(['Site', 'Year', 'Month', 'Case No.'])
            
            logger.info(f"현장별 상세 보고서 완료: {len(result)}개 레코드")
            return result
        else:
            logger.warning("현장별 상세 데이터가 없습니다")
            return pd.DataFrame()
    
    def format_excel_worksheet(self, ws, title, data):
        """엑셀 워크시트 포맷팅"""
        # 제목 추가
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", bold=True)
        
        # 데이터 추가
        if not data.empty:
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            
            # 헤더 스타일링
            for cell in ws[2]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 테두리 추가
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2, max_row=len(data)+2, min_col=1, max_col=len(data.columns)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_comprehensive_report(self):
        """종합 보고서 생성"""
        logger.info("종합 월별 보고서 생성 시작")
        
        try:
            # 데이터 로드
            df = self.load_data()
            
            # 워크북 생성
            wb = Workbook()
            
            # 기본 시트 제거
            wb.remove(wb.active)
            
            # 1. 창고별 월별 요약 시트
            warehouse_summary = self.create_warehouse_monthly_summary(df)
            if not warehouse_summary.empty:
                ws1 = wb.create_sheet("창고별_월별_요약")
                self.format_excel_worksheet(ws1, "창고별 월별 입고/출고 요약", warehouse_summary)
            
            # 2. 현장별 월별 요약 시트
            site_summary = self.create_site_monthly_summary(df)
            if not site_summary.empty:
                ws2 = wb.create_sheet("현장별_월별_요약")
                self.format_excel_worksheet(ws2, "현장별 월별 입고/누적재고 요약", site_summary)
            
            # 3. 창고별 상세 시트
            warehouse_detail = self.create_detailed_warehouse_report(df)
            if not warehouse_detail.empty:
                ws3 = wb.create_sheet("창고별_상세")
                self.format_excel_worksheet(ws3, "창고별 상세 입고/출고 내역", warehouse_detail)
            
            # 4. 현장별 상세 시트
            site_detail = self.create_detailed_site_report(df)
            if not site_detail.empty:
                ws4 = wb.create_sheet("현장별_상세")
                self.format_excel_worksheet(ws4, "현장별 상세 입고/누적재고 내역", site_detail)
            
            # 5. 데이터 개요 시트
            overview_data = pd.DataFrame({
                '항목': ['전체 레코드 수', '창고별 데이터', '현장별 데이터', '기간'],
                '값': [
                    len(df),
                    len(warehouse_summary) if not warehouse_summary.empty else 0,
                    len(site_summary) if not site_summary.empty else 0,
                    f"{df['ETD/ATD'].min().strftime('%Y-%m')} ~ {df['ETD/ATD'].max().strftime('%Y-%m')}"
                ]
            })
            
            ws5 = wb.create_sheet("데이터_개요")
            self.format_excel_worksheet(ws5, "HVDC 데이터 개요", overview_data)
            
            # 파일 저장
            wb.save(self.output_file)
            logger.info(f"종합 보고서 생성 완료: {self.output_file}")
            
            # 생성 결과 리포트
            report = {
                "생성_시간": datetime.now().isoformat(),
                "파일명": self.output_file,
                "전체_레코드": len(df),
                "창고별_요약": len(warehouse_summary) if not warehouse_summary.empty else 0,
                "현장별_요약": len(site_summary) if not site_summary.empty else 0,
                "창고별_상세": len(warehouse_detail) if not warehouse_detail.empty else 0,
                "현장별_상세": len(site_detail) if not site_detail.empty else 0,
                "시트_수": len(wb.sheetnames)
            }
            
            # 리포트 저장
            report_file = f"종합_월별_보고서_생성_리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(report_file, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            
            logger.info(f"생성 리포트 저장: {report_file}")
            
            return report
            
        except Exception as e:
            logger.error(f"보고서 생성 실패: {e}")
            raise

def main():
    """메인 실행 함수"""
    print("🏢 HVDC 종합 월별 보고서 생성 시스템")
    print("=" * 50)
    
    try:
        generator = ComprehensiveMonthlyReportGenerator()
        report = generator.generate_comprehensive_report()
        
        print(f"\n✅ 보고서 생성 완료!")
        print(f"📁 파일: {report['파일명']}")
        print(f"📊 전체 레코드: {report['전체_레코드']:,}개")
        print(f"🏭 창고별 요약: {report['창고별_요약']}개")
        print(f"🏗️ 현장별 요약: {report['현장별_요약']}개")
        print(f"📋 시트 수: {report['시트_수']}개")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        return False
    
    return True

if __name__ == "__main__":
    main() 